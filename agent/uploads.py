"""Upload manager — handles saving, validation, and metadata for user-uploaded files."""

import json
import time
import uuid
import os
import shutil
import sqlite3
from pathlib import Path
from typing import Optional
from werkzeug.utils import secure_filename

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Allowed extensions from the plan
TEXT_EXTENSIONS = {
    ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".c", ".cpp", ".h",
    ".cs", ".go", ".rs", ".rb", ".php", ".swift", ".kt", ".scala",
    ".html", ".css", ".scss", ".svg", ".xml",
    ".json", ".yaml", ".yml", ".toml", ".csv", ".sql", ".graphql",
    ".cfg", ".ini", ".env", ".sh", ".bash", ".ps1", ".bat", ".cmd",
    ".md", ".txt", ".rst", ".log",
    ".dockerfile", ".gitignore", ".editorconfig",
}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".ico"}
DOC_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".pptx"}
EXTRACTABLE_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".db", ".sqlite"}

BLOCKED_EXTENSIONS = {
    ".exe", ".dll", ".so", ".bin", ".msi", ".bat", ".cmd", ".ps1",
    ".scr", ".com", ".vbs", ".wsf", ".jar",
}

class UploadManager:
    """Manages files uploaded by the user during chat sessions."""

    def __init__(self, memory_dir: str):
        self.memory_dir = Path(memory_dir)
        self.uploads_dir = self.memory_dir / "uploads"
        self.uploads_dir.mkdir(parents=True, exist_ok=True)

    def save_upload(self, files, session_id: str) -> dict:
        """
        Save a list of files to a new upload directory.
        'files' should be a list of werkzeug FileStorage objects.
        """
        upload_id = uuid.uuid4().hex[:8]
        dest_root = self.uploads_dir / upload_id
        dest_root.mkdir(parents=True, exist_ok=True)

        manifest_files = []
        for file in files:
            raw_filename = file.filename
            if not raw_filename:
                continue

            # Handle potential path separators for folder uploads
            parts = raw_filename.replace("\\", "/").split("/")
            sanitized_parts = [secure_filename(p) for p in parts if p]
            if not sanitized_parts:
                continue
            
            relative_path = os.path.join(*sanitized_parts)
            filename = sanitized_parts[-1]
            
            ext = os.path.splitext(filename)[1].lower()
            if ext in BLOCKED_EXTENSIONS:
                continue

            file_dest = dest_root / relative_path
            file_dest.parent.mkdir(parents=True, exist_ok=True)
            
            file.save(str(file_dest))

            manifest_files.append({
                "name": filename,
                "relative_path": relative_path.replace("\\", "/"),
                "size": file_dest.stat().st_size,
                "extension": ext,
                "is_text": ext in TEXT_EXTENSIONS,
                "is_image": ext in IMAGE_EXTENSIONS,
                "is_extractable": ext in EXTRACTABLE_EXTENSIONS
            })

        manifest = {
            "upload_id": upload_id,
            "session_id": session_id,
            "created_at": time.time(),
            "files": manifest_files
        }

        manifest_path = dest_root / "manifest.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2, ensure_ascii=False)

        return manifest

    def get_upload(self, upload_id: str) -> Optional[dict]:
        """Retrieve manifest for a specific upload."""
        manifest_path = self.uploads_dir / upload_id / "manifest.json"
        if manifest_path.exists():
            try:
                with open(manifest_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return None
        return None

    def read_upload_files(self, upload_id: str) -> list:
        """Read files from an upload, extracting text from non-plain-text formats where possible."""
        manifest = self.get_upload(upload_id)
        if not manifest:
            return []

        results = []
        upload_root = self.uploads_dir / upload_id
        for f in manifest.get("files", []):
            file_path = upload_root / f["relative_path"]
            if not file_path.exists():
                continue
                
            content = None
            ext = f.get("extension", "").lower()
            
            if f.get("is_text"):
                try:
                    content = file_path.read_text(encoding="utf-8", errors="replace")
                except Exception:
                    continue
            elif ext in EXTRACTABLE_EXTENSIONS:
                content = self._extract_text(file_path, ext)
            elif f.get("is_image"):
                content = self._describe_image(file_path)
            else:
                # Fallback for binary files
                size = f.get("size", 0)
                content = f"[Binary file: {f['name']}, {size} bytes — cannot be read by this model]"

            if content:
                results.append({
                    "name": f["name"],
                    "path": f["relative_path"],
                    "content": content
                })
        return results

    def delete_upload(self, upload_id: str):
        """Delete an entire upload directory."""
        upload_path = self.uploads_dir / upload_id
        if upload_path.exists():
            shutil.rmtree(upload_path)

    def cleanup_orphans(self, max_age_hours: int = 24):
        """Remove uploads older than specified age."""
        now = time.time()
        if not self.uploads_dir.exists():
            return

        for upload_dir in self.uploads_dir.iterdir():
            if not upload_dir.is_dir():
                continue
            
            manifest_path = upload_dir / "manifest.json"
            if not manifest_path.exists():
                shutil.rmtree(upload_dir)
                continue
                
            try:
                with open(manifest_path, "r") as f:
                    manifest = json.load(f)
                created_at = manifest.get("created_at", 0)
                if now - created_at > (max_age_hours * 3600):
                    shutil.rmtree(upload_dir)
            except Exception:
                pass

    # --- Text Extraction Extractors ---

    def _extract_text(self, file_path: Path, extension: str) -> Optional[str]:
        """Dispatcher for extracting text from non-plain-text files."""
        if not file_path.exists():
            return None
            
        if extension == ".pdf":
            return self._extract_pdf(file_path)
        elif extension == ".docx":
            return self._extract_docx(file_path)
        elif extension == ".xlsx":
            return self._extract_xlsx(file_path)
        elif extension in {".db", ".sqlite"}:
            return self._extract_sqlite(file_path)
        return None

    def _extract_pdf(self, file_path: Path) -> str:
        if not pdfplumber:
            return f"[PDF file: {file_path.name} — install 'pdfplumber' to extract text]"
        try:
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                    if len(text) > 50000:
                        break
            return self._truncate(text)
        except Exception as e:
            return f"[Error extracting PDF {file_path.name}: {str(e)}]"

    def _extract_docx(self, file_path: Path) -> str:
        if not Document:
            return f"[DOCX file: {file_path.name} — install 'python-docx' to extract text]"
        try:
            doc = Document(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            return self._truncate(text)
        except Exception as e:
            return f"[Error extracting DOCX {file_path.name}: {str(e)}]"

    def _extract_xlsx(self, file_path: Path) -> str:
        if not openpyxl:
            return f"[XLSX file: {file_path.name} — install 'openpyxl' to extract text]"
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            output = []
            for sheet_name in wb.sheetnames[:5]: # Limit to 5 sheets
                sheet = wb[sheet_name]
                output.append(f"--- Sheet: {sheet_name} ---")
                for row in sheet.iter_rows(max_row=200, values_only=True): # Limit to 200 rows
                    if any(row):
                        output.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
                if len("\n".join(output)) > 50000:
                    break
            return self._truncate("\n".join(output))
        except Exception as e:
            return f"[Error extracting XLSX {file_path.name}: {str(e)}]"

    def _extract_sqlite(self, file_path: Path) -> str:
        try:
            conn = sqlite3.connect(file_path)
            cursor = conn.cursor()
            output = [f"--- SQLite Database: {file_path.name} ---"]
            
            # Get tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]
            
            for table in tables:
                # Schema
                cursor.execute(f"SELECT sql FROM sqlite_master WHERE type='table' AND name='{table}';")
                schema = cursor.fetchall()[0][0]
                output.append(f"\nTable: {table}\nSchema: {schema}")
                
                # Data (first 50 rows)
                cursor.execute(f"SELECT * FROM {table} LIMIT 50;")
                rows = cursor.fetchall()
                if rows:
                    output.append("Data (up to 50 rows):")
                    for row in rows:
                        output.append(str(row))
                
                if len("\n".join(output)) > 50000:
                    break
                    
            conn.close()
            return self._truncate("\n".join(output))
        except Exception as e:
            return f"[Error reading SQLite {file_path.name}: {str(e)}]"

    def _describe_image(self, file_path: Path) -> str:
        size_kb = file_path.stat().st_size / 1024
        return f"[Image: {file_path.name}, {size_kb:.1f} KB — image content cannot be analyzed by this model]"

    def _truncate(self, text: str, limit: int = 50000) -> str:
        if len(text) > limit:
            return text[:limit] + f"\n\n... [Content truncated at {limit} characters] ..."
        return text
